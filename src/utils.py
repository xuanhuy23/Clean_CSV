import os
import pandas as pd
import numpy as np
import shutil
import tempfile
import time
import logging
import json
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
from datetime import datetime
import openpyxl

def save_to_excel(df: pd.DataFrame, output_path: str, sheet_name: str = 'Dữ liệu', index: bool = False) -> bool:
    """
    Lưu DataFrame vào file Excel.
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần lưu
    output_path : str
        Đường dẫn đến file Excel đầu ra
    sheet_name : str, default='Dữ liệu'
        Tên sheet cho dữ liệu
    index : bool, default=False
        Có lưu chỉ số hay không
        
    Returns:
    --------
    bool
        True nếu lưu thành công, False nếu thất bại
    """
    # Kiểm tra DataFrame
    if df is None or df.empty:
        logging.error("DataFrame trống không thể lưu vào Excel")
        return False
        
    try:
        # Đảm bảo thư mục đầu ra tồn tại
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir, exist_ok=True)
            except PermissionError:
                logging.error(f"Không thể tạo thư mục '{output_dir}', lỗi phân quyền")
                return False
                
        # Kiểm tra xem có engine Excel phù hợp không
        try:
            import importlib
            if importlib.util.find_spec("openpyxl") is None and importlib.util.find_spec("xlsxwriter") is None:
                logging.warning("Không có engine Excel nào được cài đặt. Vui lòng cài đặt 'openpyxl' hoặc 'xlsxwriter'")
                logging.warning("Đang thử lưu mà không chỉ định engine...")
                df.to_excel(output_path, sheet_name=sheet_name, index=index)
            else:
                engine = "openpyxl" if importlib.util.find_spec("openpyxl") is not None else "xlsxwriter"
                logging.info(f"Đang sử dụng engine Excel: {engine}")
                df.to_excel(output_path, sheet_name=sheet_name, index=index, engine=engine)
                
            logging.info(f"Đã lưu dữ liệu thành công vào file Excel: {output_path}")
            return True
        except ImportError:
            # Fallback nếu cả hai engine đều không khả dụng
            try:
                df.to_excel(output_path, sheet_name=sheet_name, index=index)
                logging.info(f"Đã lưu dữ liệu thành công vào file Excel: {output_path}")
                return True
            except Exception as e:
                logging.error(f"Lỗi khi lưu Excel (không có engine): {str(e)}")
                return False
                
    except PermissionError:
        logging.error(f"Không thể lưu file '{output_path}', kiểm tra quyền truy cập")
        return False
    except Exception as e:
        logging.error(f"Lỗi khi lưu DataFrame vào Excel: {str(e)}")
        # Thử các phương pháp thay thế để lưu Excel
        try:
            logging.info("Đang thử phương pháp thay thế để lưu Excel...")
            with pd.ExcelWriter(output_path, engine='openpyxl' if importlib.util.find_spec("openpyxl") is not None else 'xlsxwriter') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
            logging.info(f"Đã lưu dữ liệu thành công vào file Excel (phương pháp thay thế): {output_path}")
            return True
        except Exception as ex:
            logging.error(f"Phương pháp thay thế cũng thất bại: {str(ex)}")
            return False

def save_data(df: pd.DataFrame, 
              output_path: str, 
              index: bool = False, 
              encoding: str = 'utf-8', 
              sep: str = ',',
              excel: bool = False,
              sheet_name: str = 'Data',
              save_metadata: bool = True,
              **kwargs) -> bool:
    """
    Lưu DataFrame ra file với xử lý lỗi nâng cao.
    Hỗ trợ cả định dạng CSV và Excel.
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần lưu
    output_path : str
        Đường dẫn đến file đầu ra
    index : bool, default False
        Có lưu chỉ mục DataFrame hay không
    encoding : str, default 'utf-8'
        Mã hóa ký tự file (chỉ áp dụng cho CSV)
    sep : str, default ','
        Ký tự phân cách cột (chỉ áp dụng cho CSV)
    excel : bool, default False
        Nếu True, lưu dưới định dạng Excel (.xlsx) thay vì CSV
    sheet_name : str, default 'Data'
        Tên sheet trong file Excel (chỉ áp dụng khi excel=True)
    save_metadata : bool, default True
        Lưu thông tin metadata (kiểu dữ liệu, định dạng) vào file đi kèm
    **kwargs : dict
        Các tham số khác cho pd.to_csv() hoặc pd.to_excel()
        
    Returns:
    --------
    bool
        True nếu lưu thành công, False nếu có lỗi
    """
    # Kiểm tra định dạng file đầu ra
    if os.path.splitext(output_path)[1].lower() in ['.xlsx', '.xls']:
        return save_to_excel(
            df, 
            output_path, 
            sheet_name=sheet_name, 
            index=index, 
            **kwargs
        )
    else:
        # Kiểm tra DataFrame
        if df is None:
            logging.error("Không thể lưu: DataFrame là None")
            print("Lỗi: DataFrame là None, không thể lưu file")
            return False
    
    if df.empty:
        logging.warning("DataFrame trống, lưu file trống")
        print("Cảnh báo: DataFrame trống, sẽ lưu file trống")
    
    # Chuẩn hóa đường dẫn
    output_path = os.path.normpath(output_path)
    output_dir = os.path.dirname(output_path)
    
    # Đảm bảo thư mục đầu ra tồn tại
    if output_dir and not os.path.exists(output_dir):
        try:
            print(f"Tạo thư mục đầu ra: {output_dir}")
            Path(output_dir).mkdir(parents=True, exist_ok=True)
            logging.info(f"Đã tạo thư mục đầu ra: {output_dir}")
        except PermissionError:
            logging.error(f"Lỗi quyền truy cập: Không thể tạo thư mục {output_dir}")
            print(f"Lỗi: Không có quyền tạo thư mục {output_dir}")
            
            # Thử lưu vào thư mục tạm
            try:
                temp_dir = tempfile.gettempdir()
                temp_output = os.path.join(temp_dir, os.path.basename(output_path))
                logging.warning(f"Thử lưu vào thư mục tạm: {temp_output}")
                print(f"Thử lưu vào thư mục tạm: {temp_output}")
                output_path = temp_output
            except Exception as e2:
                logging.error(f"Không thể tạo thư mục tạm: {str(e2)}")
                print(f"Lỗi: Không thể tạo thư mục tạm: {str(e2)}")
                
                # Thử lưu vào thư mục home của người dùng
                try:
                    home_dir = os.path.expanduser("~")
                    home_output = os.path.join(home_dir, os.path.basename(output_path))
                    logging.warning(f"Thử lưu vào thư mục home: {home_output}")
                    print(f"Thử lưu vào thư mục home: {home_output}")
                    output_path = home_output
                except Exception as e3:
                    logging.error(f"Không thể sử dụng thư mục home: {str(e3)}")
                    print(f"Lỗi: Không thể sử dụng thư mục home: {str(e3)}")
                    return False
        except Exception as e:
            logging.error(f"Lỗi khi tạo thư mục đầu ra {output_dir}: {str(e)}")
            print(f"Lỗi: Không thể tạo thư mục đầu ra {output_dir}: {str(e)}")
            return False
    
    # Thử lưu DataFrame với cơ chế thử lại
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            # Lưu DataFrame
            print(f"Đang lưu DataFrame ({df.shape[0]} dòng x {df.shape[1]} cột) vào {output_path}")
            logging.info(f"Đang lưu DataFrame ({df.shape[0]} dòng x {df.shape[1]} cột) vào {output_path}")
            df.to_csv(output_path, index=index, encoding=encoding, sep=sep, **kwargs)
            
            # Kiểm tra xem file có tồn tại sau khi lưu không
            if not os.path.exists(output_path):
                logging.error(f"File không tồn tại sau khi lưu: {output_path}")
                print(f"Lỗi: File không tồn tại sau khi lưu: {output_path}")
                continue
                
            # Kiểm tra kích thước file
            file_size = os.path.getsize(output_path)
            if file_size == 0 and not df.empty:
                logging.error(f"File đầu ra có kích thước 0: {output_path}")
                print(f"Lỗi: File đầu ra có kích thước 0: {output_path}")
                continue
                
            if file_size < 100 and not df.empty and df.shape[0] > 5:  # Kích thước quá nhỏ
                logging.warning(f"File đầu ra có kích thước đáng ngờ ({file_size} bytes): {output_path}")
                print(f"Cảnh báo: File đầu ra có kích thước đáng ngờ ({file_size} bytes): {output_path}")
            
            # Thử đọc file để xác minh nó có thể được đọc
            try:
                test_df = pd.read_csv(output_path, nrows=5, encoding=encoding, sep=sep)
                if test_df is not None:
                    logging.info(f"Đã xác minh file có thể đọc lại được: {output_path}")
            except Exception as e:
                logging.warning(f"Không thể đọc lại file (nhưng file đã được tạo): {str(e)}")
                print(f"Cảnh báo: Không thể đọc lại file đã lưu: {str(e)}")
                
                # Lưu metadata về kiểu dữ liệu nếu được yêu cầu
                if save_metadata:
                    try:
                        metadata_path = f"{os.path.splitext(output_path)[0]}.meta.json"
                        metadata = {
                            'created_at': datetime.now().isoformat(),
                            'rows': df.shape[0],
                            'columns': df.shape[1],
                            'column_types': {col: str(dtype) for col, dtype in df.dtypes.items()},
                            'format_hints': {}
                        }
                        
                        # Thêm gợi ý định dạng cho từng cột
                        for col in df.columns:
                            if pd.api.types.is_numeric_dtype(df[col]):
                                if pd.api.types.is_integer_dtype(df[col]):
                                    metadata['format_hints'][col] = 'integer'
                                else:
                                    metadata['format_hints'][col] = 'float'
                                    # Xác định số chữ số thập phân phổ biến nhất
                                    decimal_places = 2  # Mặc định
                                    try:
                                        non_na_values = df[col].dropna()
                                        if len(non_na_values) > 0:
                                            # Tính số chữ số thập phân phổ biến
                                            decimal_counts = non_na_values.astype(str).str.extract(r'\.(\d+)')[0].str.len().value_counts()
                                            if not decimal_counts.empty:
                                                decimal_places = decimal_counts.index[0]
                                    except:
                                        pass
                                    metadata['format_hints'][col] = f'float:{decimal_places}'
                            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                                metadata['format_hints'][col] = 'datetime'
                            elif pd.api.types.is_categorical_dtype(df[col]):
                                metadata['format_hints'][col] = 'category'
                            else:
                                metadata['format_hints'][col] = 'text'
                        
                        # Lưu metadata ra file json
                        with open(metadata_path, 'w', encoding='utf-8') as f:
                            json.dump(metadata, f, ensure_ascii=False, indent=2)
                        
                        logging.info(f"Đã lưu metadata: {metadata_path}")
                        print(f"Đã lưu thông tin kiểu dữ liệu vào: {metadata_path}")
                    except Exception as e:
                        logging.warning(f"Không thể lưu metadata: {str(e)}")
                        print(f"Cảnh báo: Không thể lưu metadata: {str(e)}")
                
            logging.info(f"Đã lưu thành công: {output_path} ({file_size:,} bytes)")
            print(f"Đã lưu dữ liệu đã làm sạch vào: {output_path} ({file_size:,} bytes)")
            return True
            
        except UnicodeEncodeError:
            # Thử với một encoding khác
            alt_encoding = 'latin1' if encoding != 'latin1' else 'utf-8'
            logging.warning(f"Lỗi mã hóa. Thử với {alt_encoding}")
            print(f"Lỗi mã hóa. Thử lưu với encoding khác: {alt_encoding}")
            try:
                df.to_csv(output_path, index=index, encoding=alt_encoding, sep=sep, **kwargs)
                logging.info(f"Đã lưu với encoding thay thế {alt_encoding}: {output_path}")
                print(f"Đã lưu dữ liệu với encoding thay thế {alt_encoding}: {output_path}")
                return True
            except Exception as e2:
                logging.error(f"Lỗi khi lưu với encoding thay thế: {str(e2)}")
                print(f"Lỗi khi lưu với encoding thay thế: {str(e2)}")
                
        except PermissionError:
            logging.error(f"Lỗi quyền truy cập khi lưu file: {output_path}")
            print(f"Lỗi quyền truy cập khi lưu file: {output_path}")
            # Thử lưu vào thư mục tạm
            try:
                temp_path = os.path.join(tempfile.gettempdir(), os.path.basename(output_path))
                print(f"Thử lưu vào đường dẫn tạm: {temp_path}")
                df.to_csv(temp_path, index=index, encoding=encoding, sep=sep, **kwargs)
                logging.warning(f"Đã lưu vào đường dẫn tạm do lỗi quyền truy cập: {temp_path}")
                print(f"File đã được lưu vào vị trí tạm (do lỗi quyền truy cập): {temp_path}")
                return True
            except Exception as e2:
                logging.error(f"Không thể lưu vào đường dẫn tạm: {str(e2)}")
                print(f"Không thể lưu vào đường dẫn tạm: {str(e2)}")
                
        except OSError as e:
            # Kiểm tra xem có đủ dung lượng ổ đĩa không
            try:
                free_space = shutil.disk_usage(os.path.dirname(output_path) or '.').free
                required_space = df.memory_usage(deep=True).sum() * 1.5  # Ước tính
                
                if free_space < required_space:
                    logging.error(f"Không đủ dung lượng ổ đĩa. Cần {required_space/1024/1024:.2f}MB, " 
                                 f"còn trống {free_space/1024/1024:.2f}MB")
                    print(f"Lỗi: Không đủ dung lượng ổ đĩa để lưu file.")
                    return False
            except Exception:
                pass  # Bỏ qua lỗi khi kiểm tra dung lượng
                
            logging.error(f"Lỗi hệ thống khi lưu: {str(e)}")
            print(f"Lỗi hệ thống khi lưu: {str(e)}")
            
        except Exception as e:
            logging.error(f"Lỗi khi lưu file (lần thử {attempt}/{max_retries}): {str(e)}")
            print(f"Lỗi khi lưu file (lần thử {attempt}/{max_retries}): {str(e)}")
            
        # Chờ một chút trước khi thử lại
        if attempt < max_retries:
            wait_time = 2 ** attempt  # Tăng thời gian chờ theo cấp số nhân
            logging.info(f"Chờ {wait_time} giây trước khi thử lại...")
            print(f"Chờ {wait_time} giây trước khi thử lại...")
            time.sleep(wait_time)
    
    logging.error(f"Không thể lưu file sau {max_retries} lần thử")
    print(f"Lỗi: Không thể lưu file sau nhiều lần thử. Kiểm tra logs để biết thêm chi tiết.")
    return False

def get_summary_stats(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Trả về thống kê tóm tắt về DataFrame
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần phân tích
        
    Returns:
    --------
    Dict[str, Any]
        Từ điển chứa các thống kê tóm tắt
    """
    summary = {
        'row_count': df.shape[0],
        'column_count': df.shape[1],
        'memory_usage': df.memory_usage(deep=True).sum(),
        'missing_values': df.isnull().sum().sum(),
        'dtypes': dict(df.dtypes.value_counts()),
        'columns': df.columns.tolist()
    }
    
    # Tính tỷ lệ giá trị thiếu
    if df.shape[0] > 0 and df.shape[1] > 0:
        summary['missing_percent'] = (summary['missing_values'] / (df.shape[0] * df.shape[1])) * 100
    else:
        summary['missing_percent'] = 0
    
    return summary

def identify_categorical_columns(df: pd.DataFrame, threshold: float = 0.05) -> List[str]:
    """
    Xác định các cột có thể là dạng phân loại (categorical)
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần kiểm tra
    threshold : float, default 0.05
        Ngưỡng tỷ lệ giá trị duy nhất/tổng số dòng để xác định cột phân loại
    
    Returns:
    --------
    List[str]
        Danh sách các cột được xác định là dạng phân loại
    """
    categorical_cols = []
    
    for col in df.columns:
        # Bỏ qua các cột đã là kiểu category
        if pd.api.types.is_categorical_dtype(df[col]):
            categorical_cols.append(col)
            continue
        
        # Kiểm tra tỷ lệ giá trị duy nhất
        n_unique = df[col].nunique()
        n_rows = len(df)
        
        # Nếu tỷ lệ giá trị duy nhất thấp hơn ngưỡng, có thể là cột phân loại
        if n_unique / n_rows <= threshold:
            categorical_cols.append(col)
    
    return categorical_cols

def identify_potential_datetime_columns(df: pd.DataFrame) -> List[str]:
    """
    Xác định các cột có thể chứa dữ liệu ngày tháng
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần kiểm tra
    
    Returns:
    --------
    List[str]
        Danh sách các cột có thể chứa dữ liệu ngày tháng
    """
    datetime_cols = []
    date_keywords = ['date', 'time', 'day', 'month', 'year', 'ngay', 'thang', 'nam']
    
    for col in df.columns:
        # Đã là kiểu datetime
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            datetime_cols.append(col)
            continue
        
        # Kiểm tra tên cột
        if any(keyword in col.lower() for keyword in date_keywords):
            datetime_cols.append(col)
            continue
        
        # Nếu là kiểu chuỗi, thử phát hiện dựa trên mẫu
        if df[col].dtype == 'object':
            sample = df[col].dropna().head(5).astype(str).tolist()
            if sample and all(
                any(c in s for c in ['-', '/', '.']) and 
                sum(c.isdigit() for c in s) >= 4  # Ít nhất có 4 chữ số
                for s in sample
            ):
                datetime_cols.append(col)
    
    return datetime_cols

def get_column_quality_score(df: pd.DataFrame, col: str) -> Dict[str, Any]:
    """
    Tính điểm chất lượng cho một cột dữ liệu
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame chứa cột cần kiểm tra
    col : str
        Tên cột cần kiểm tra
    
    Returns:
    --------
    Dict[str, Any]
        Điểm chất lượng và thông tin chi tiết
    """
    if col not in df.columns:
        return {'score': 0, 'issues': ['Cột không tồn tại']}
    
    issues = []
    score = 100  # Bắt đầu với điểm tối đa
    
    # Kiểm tra giá trị thiếu
    missing_count = df[col].isnull().sum()
    missing_percent = (missing_count / len(df)) * 100
    if missing_percent > 0:
        issues.append(f"Có {missing_count} giá trị thiếu ({missing_percent:.2f}%)")
        score -= min(50, missing_percent)  # Giảm điểm tùy theo % giá trị thiếu
    
    # Kiểm tra kiểu dữ liệu và các vấn đề khác tùy theo kiểu
    if pd.api.types.is_numeric_dtype(df[col]):
        # Kiểm tra outliers cho cột số
        q1 = df[col].quantile(0.25)
        q3 = df[col].quantile(0.75)
        iqr = q3 - q1
        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        outliers = ((df[col] < lower_bound) | (df[col] > upper_bound)).sum()
        outlier_percent = (outliers / len(df)) * 100
        
        if outlier_percent > 0:
            issues.append(f"Có {outliers} giá trị ngoại lai ({outlier_percent:.2f}%)")
            score -= min(30, outlier_percent)
    
    elif df[col].dtype == 'object':
        # Kiểm tra tính nhất quán của chuỗi
        if df[col].str.strip().ne(df[col]).any():
            issues.append("Có khoảng trắng thừa")
            score -= 10
        
        # Kiểm tra case không nhất quán
        if not df[col].str.islower().all() and not df[col].str.isupper().all():
            issues.append("Chữ hoa/thường không nhất quán")
            score -= 10
    
    # Đánh giá chung
    if score >= 90:
        quality = "Rất tốt"
    elif score >= 75:
        quality = "Tốt"
    elif score >= 50:
        quality = "Trung bình"
    elif score >= 25:
        quality = "Kém"
    else:
        quality = "Rất kém"
    
    return {
        'score': score,
        'quality': quality,
        'issues': issues
    }

def generate_html_report(report_data: Dict[str, Any], output_path: str, include_plots: bool = True) -> bool:
    """
    Tạo báo cáo HTML tóm tắt quá trình làm sạch dữ liệu.
    
    Parameters:
    -----------
    report_data : Dict[str, Any]
        Dữ liệu báo cáo, bao gồm thông tin về các bước xử lý, thống kê, v.v.
    output_path : str
        Đường dẫn đến file HTML đầu ra
    include_plots : bool, default True
        Có tạo và đưa biểu đồ vào báo cáo hay không
        
    Returns:
    --------
    bool
        True nếu tạo báo cáo thành công, False nếu có lỗi
    """
    try:
        # Đảm bảo thư mục đầu ra tồn tại
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # Tên file báo cáo
        report_filename = os.path.basename(output_path)
        
        # Tạo thư mục cho đồ thị
        plots_dir = os.path.join(output_dir, "plots")
        os.makedirs(plots_dir, exist_ok=True)
        
        # Đường dẫn tương đối đến thư mục đồ thị
        plots_relative_path = os.path.join("plots")
        
        # Danh sách các đường dẫn đến đồ thị
        plot_paths = []
        
        # Tạo các đồ thị nếu được yêu cầu
        if include_plots and 'stats' in report_data:
            try:
                # Đồ thị so sánh kích thước dữ liệu trước và sau khi làm sạch
                if 'before' in report_data['stats'] and 'after' in report_data['stats']:
                    before = report_data['stats']['before']
                    after = report_data['stats']['after']
                    
                    if 'row_count' in before and 'row_count' in after:
                        # Biểu đồ so sánh số lượng dòng
                        plt.figure(figsize=(10, 6))
                        labels = ['Trước khi làm sạch', 'Sau khi làm sạch']
                        row_counts = [before['row_count'], after['row_count']]
                        
                        plt.bar(labels, row_counts, color=['#5DA5DA', '#4D4D4D'])
                        plt.title('So sánh số lượng dòng trước và sau khi làm sạch', fontsize=14)
                        plt.ylabel('Số lượng dòng')
                        plt.grid(axis='y', linestyle='--', alpha=0.7)
                        
                        # Thêm giá trị lên các cột
                        for i, v in enumerate(row_counts):
                            plt.text(i, v + 0.1, str(v), ha='center', fontsize=12)
                        
                        # Lưu biểu đồ
                        row_count_plot = os.path.join(plots_dir, 'row_count_comparison.png')
                        plt.tight_layout()
                        plt.savefig(row_count_plot)
                        plt.close()
                        
                        plot_paths.append({
                            'title': 'So sánh số lượng dòng',
                            'path': os.path.join(plots_relative_path, 'row_count_comparison.png')
                        })
                    
                    if 'missing_values' in before and 'missing_values' in after:
                        # Biểu đồ so sánh số lượng giá trị thiếu
                        plt.figure(figsize=(10, 6))
                        labels = ['Trước khi làm sạch', 'Sau khi làm sạch']
                        missing_values = [before['missing_values'], after['missing_values']]
                        
                        plt.bar(labels, missing_values, color=['#F15854', '#60BD68'])
                        plt.title('So sánh số lượng giá trị thiếu trước và sau khi làm sạch', fontsize=14)
                        plt.ylabel('Số lượng giá trị thiếu')
                        plt.grid(axis='y', linestyle='--', alpha=0.7)
                        
                        # Thêm giá trị lên các cột
                        for i, v in enumerate(missing_values):
                            plt.text(i, v + 0.1, str(v), ha='center', fontsize=12)
                        
                        # Lưu biểu đồ
                        missing_values_plot = os.path.join(plots_dir, 'missing_values_comparison.png')
                        plt.tight_layout()
                        plt.savefig(missing_values_plot)
                        plt.close()
                        
                        plot_paths.append({
                            'title': 'So sánh số lượng giá trị thiếu',
                            'path': os.path.join(plots_relative_path, 'missing_values_comparison.png')
                        })
                    
                    if 'dtypes' in before and 'dtypes' in after:
                        # Biểu đồ so sánh kiểu dữ liệu
                        plt.figure(figsize=(12, 6))
                        
                        # Tính số lượng cột theo kiểu dữ liệu
                        before_dtypes = {}
                        for dtype, count in before['dtypes'].items():
                            dtype_name = str(dtype).split('(')[0]
                            before_dtypes[dtype_name] = before_dtypes.get(dtype_name, 0) + count
                        
                        after_dtypes = {}
                        for dtype, count in after['dtypes'].items():
                            dtype_name = str(dtype).split('(')[0]
                            after_dtypes[dtype_name] = after_dtypes.get(dtype_name, 0) + count
                        
                        # Gộp tất cả các kiểu dữ liệu
                        all_dtypes = list(set(list(before_dtypes.keys()) + list(after_dtypes.keys())))
                        all_dtypes.sort()
                        
                        # Chuẩn bị dữ liệu cho biểu đồ
                        before_counts = [before_dtypes.get(dtype, 0) for dtype in all_dtypes]
                        after_counts = [after_dtypes.get(dtype, 0) for dtype in all_dtypes]
                        
                        x = range(len(all_dtypes))
                        width = 0.35
                        
                        fig, ax = plt.subplots(figsize=(12, 6))
                        rects1 = ax.bar([i - width/2 for i in x], before_counts, width, label='Trước khi làm sạch')
                        rects2 = ax.bar([i + width/2 for i in x], after_counts, width, label='Sau khi làm sạch')
                        
                        ax.set_ylabel('Số lượng cột')
                        ax.set_title('So sánh số lượng cột theo kiểu dữ liệu')
                        ax.set_xticks(x)
                        ax.set_xticklabels(all_dtypes)
                        ax.legend()
                        
                        # Lưu biểu đồ
                        dtypes_plot = os.path.join(plots_dir, 'dtypes_comparison.png')
                        plt.tight_layout()
                        plt.savefig(dtypes_plot)
                        plt.close()
                        
                        plot_paths.append({
                            'title': 'So sánh kiểu dữ liệu',
                            'path': os.path.join(plots_relative_path, 'dtypes_comparison.png')
                        })
            except Exception as e:
                logging.error(f"Lỗi khi tạo đồ thị cho báo cáo: {str(e)}")
                print(f"Lỗi khi tạo đồ thị cho báo cáo: {str(e)}")
        
        # Bắt đầu tạo nội dung HTML
        html_content = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Báo cáo làm sạch dữ liệu - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            color: #333;
        }}
        h1, h2, h3, h4 {{
            color: #2c3e50;
        }}
        .container {{
            margin-bottom: 30px;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 20px 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        th {{
            background-color: #f2f2f2;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .step {{
            background-color: #f8f9fa;
            border-left: 4px solid #4CAF50;
            padding: 15px;
            margin: 15px 0;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .summary-card {{
            background-color: #fff;
            border-radius: 8px;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            padding: 20px;
        }}
        .metric {{
            font-size: 24px;
            font-weight: bold;
            margin: 10px 0;
        }}
        .metric-label {{
            font-size: 14px;
            color: #666;
        }}
        .plots-container {{
            display: flex;
            flex-wrap: wrap;
            gap: 20px;
            margin: 20px 0;
        }}
        .plot-card {{
            background-color: #fff;
            border-radius: 8px;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            padding: 15px;
            flex: 1 1 300px;
            max-width: 100%;
        }}
        .plot-image {{
            max-width: 100%;
            height: auto;
        }}
        .header {{
            background-color: #2c3e50;
            color: white;
            padding: 20px;
            margin-bottom: 20px;
            border-radius: 8px;
        }}
        .header h1 {{
            color: white;
            margin: 0;
        }}
        .validation-result {{
            border-left: 4px solid #3498db;
            padding: 15px;
            margin: 15px 0;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Báo cáo làm sạch dữ liệu</h1>
        <p>Thời gian tạo: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>

    <div class="container">
        <h2>Thông tin chung</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <div class="metric-label">File đầu vào</div>
                <div class="metric">{os.path.basename(report_data.get('input_file', 'N/A'))}</div>
            </div>
            <div class="summary-card">
                <div class="metric-label">File đầu ra</div>
                <div class="metric">{os.path.basename(report_data.get('output_file', 'N/A'))}</div>
            </div>
        </div>
    </div>
"""

        # Thêm phần thống kê
        if 'stats' in report_data and 'before' in report_data['stats'] and 'after' in report_data['stats']:
            before = report_data['stats']['before']
            after = report_data['stats']['after']
            
            html_content += f"""
    <div class="container">
        <h2>Thống kê dữ liệu</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <div class="metric-label">Số dòng ban đầu</div>
                <div class="metric">{before.get('row_count', 'N/A')}</div>
            </div>
            <div class="summary-card">
                <div class="metric-label">Số dòng sau khi làm sạch</div>
                <div class="metric">{after.get('row_count', 'N/A')}</div>
            </div>
            <div class="summary-card">
                <div class="metric-label">Số cột</div>
                <div class="metric">{after.get('column_count', 'N/A')}</div>
            </div>
            <div class="summary-card">
                <div class="metric-label">Giá trị thiếu ban đầu</div>
                <div class="metric">{before.get('missing_values', 'N/A')}</div>
            </div>
            <div class="summary-card">
                <div class="metric-label">Giá trị thiếu sau khi làm sạch</div>
                <div class="metric">{after.get('missing_values', 'N/A')}</div>
            </div>
            <div class="summary-card">
                <div class="metric-label">Tỷ lệ giá trị thiếu sau khi làm sạch</div>
                <div class="metric">{after.get('missing_percent', 'N/A'):.2f}%</div>
            </div>
        </div>
    </div>
"""

        # Thêm phần các bước xử lý
        if 'processing_steps' in report_data and report_data['processing_steps']:
            html_content += f"""
    <div class="container">
        <h2>Các bước xử lý dữ liệu</h2>
"""
            for i, step in enumerate(report_data['processing_steps']):
                html_content += f"""
        <div class="step">
            <h3>Bước {i+1}: {step['step']}</h3>
            <p><strong>Phương pháp:</strong> {step['method']}</p>
            <p><strong>Chi tiết:</strong> {step.get('details', 'Không có thông tin chi tiết')}</p>
"""
                # Nếu có thông tin về các cột đã xử lý
                if 'columns' in step and step['columns']:
                    html_content += f"""
            <p><strong>Các cột đã xử lý:</strong> {', '.join(step['columns'])}</p>
"""
                html_content += """
        </div>
"""
            html_content += """
    </div>
"""

        # Thêm phần các vấn đề dữ liệu
        if 'data_issues' in report_data:
            issues = report_data['data_issues']
            html_content += f"""
    <div class="container">
        <h2>Các vấn đề dữ liệu đã phát hiện</h2>
"""
            if 'duplicate_rows' in issues:
                html_content += f"""
        <div class="validation-result">
            <h3>Dòng trùng lặp</h3>
            <p>Số lượng dòng trùng lặp: {issues['duplicate_rows']}</p>
        </div>
"""
            
            if 'missing_values' in issues and issues['missing_values']:
                html_content += f"""
        <div class="validation-result">
            <h3>Giá trị thiếu</h3>
            <table>
                <tr>
                    <th>Cột</th>
                    <th>Số lượng giá trị thiếu</th>
                </tr>
"""
                for col, count in issues['missing_values'].items():
                    html_content += f"""
                <tr>
                    <td>{col}</td>
                    <td>{count}</td>
                </tr>
"""
                html_content += """
            </table>
        </div>
"""
            
            if 'inconsistent_strings' in issues and issues['inconsistent_strings']:
                html_content += f"""
        <div class="validation-result">
            <h3>Chuỗi không nhất quán</h3>
            <table>
                <tr>
                    <th>Cột</th>
                    <th>Số lượng giá trị không nhất quán</th>
                </tr>
"""
                for col, count in issues['inconsistent_strings'].items():
                    html_content += f"""
                <tr>
                    <td>{col}</td>
                    <td>{count}</td>
                </tr>
"""
                html_content += """
            </table>
        </div>
"""
            
            html_content += """
    </div>
"""

        # Thêm phần kết quả xác thực
        if 'validation_results' in report_data:
            validation = report_data['validation_results']
            html_content += f"""
    <div class="container">
        <h2>Kết quả xác thực dữ liệu</h2>
"""
            
            if 'shape' in validation:
                shape = validation['shape']
                html_content += f"""
        <div class="validation-result">
            <h3>Kích thước dữ liệu</h3>
            <p><strong>Ban đầu:</strong> {shape['original'][0]} dòng x {shape['original'][1]} cột</p>
            <p><strong>Sau khi làm sạch:</strong> {shape['cleaned'][0]} dòng x {shape['cleaned'][1]} cột</p>
            <p><strong>Thay đổi:</strong> {shape['difference'][0]} dòng, {shape['difference'][1]} cột</p>
        </div>
"""
            
            if 'missing_values' in validation and 'original_total' in validation['missing_values'] and 'cleaned_total' in validation['missing_values']:
                html_content += f"""
        <div class="validation-result">
            <h3>Giá trị thiếu</h3>
            <p><strong>Tổng số giá trị thiếu ban đầu:</strong> {validation['missing_values']['original_total']}</p>
            <p><strong>Tổng số giá trị thiếu sau làm sạch:</strong> {validation['missing_values']['cleaned_total']}</p>
        </div>
"""
            
            html_content += """
    </div>
"""

        # Thêm phần đồ thị
        if plot_paths:
            html_content += f"""
    <div class="container">
        <h2>Biểu đồ</h2>
        <div class="plots-container">
"""
            for plot in plot_paths:
                html_content += f"""
            <div class="plot-card">
                <h3>{plot['title']}</h3>
                <img src="{plot['path']}" alt="{plot['title']}" class="plot-image">
            </div>
"""
            html_content += """
        </div>
    </div>
"""

        # Kết thúc HTML
        html_content += """
</body>
</html>
"""

        # Ghi nội dung HTML ra file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"Đã tạo báo cáo HTML tại: {output_path}")
        logging.info(f"Đã tạo báo cáo HTML tại: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo HTML: {str(e)}")
        logging.error(f"Lỗi khi tạo báo cáo HTML: {str(e)}")
        return False

def process_dataframe_in_batches(
    df: pd.DataFrame, 
    process_func: callable, 
    batch_size: int = 5000, 
    *args, 
    **kwargs
) -> pd.DataFrame:
    """
    Xử lý một DataFrame theo lô (batch) để tối ưu hóa bộ nhớ
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần xử lý
    process_func : callable
        Hàm xử lý dữ liệu, nhận vào một batch của DataFrame và trả về DataFrame đã xử lý
    batch_size : int, default 5000
        Kích thước của mỗi lô
    *args, **kwargs:
        Tham số bổ sung cho hàm xử lý
        
    Returns:
    --------
    pd.DataFrame
        DataFrame đã được xử lý theo lô
    """
    if df is None or df.empty:
        logging.warning("DataFrame trống hoặc None, không thể xử lý theo lô")
        return df
        
    if len(df) <= batch_size:
        # Nếu kích thước DataFrame nhỏ hơn hoặc bằng batch_size, xử lý trực tiếp
        logging.info(f"Xử lý toàn bộ DataFrame ({len(df)} dòng) trong một lô duy nhất")
        return process_func(df, *args, **kwargs)
    
    # Chia DataFrame thành các lô
    num_batches = (len(df) // batch_size) + (1 if len(df) % batch_size > 0 else 0)
    logging.info(f"Bắt đầu xử lý DataFrame ({len(df)} dòng) trong {num_batches} lô, mỗi lô {batch_size} dòng")
    
    result_dfs = []
    for i in range(num_batches):
        start_idx = i * batch_size
        end_idx = min((i + 1) * batch_size, len(df))
        
        # Trích xuất lô dữ liệu
        batch_df = df.iloc[start_idx:end_idx].copy()
        
        # Áp dụng hàm xử lý cho lô
        logging.info(f"Xử lý lô {i+1}/{num_batches}: dòng {start_idx} đến {end_idx-1}")
        processed_batch = process_func(batch_df, *args, **kwargs)
        
        # Lưu kết quả
        result_dfs.append(processed_batch)
        
        # Giải phóng bộ nhớ
        del batch_df
        
        # Gọi garbage collector để đảm bảo bộ nhớ được giải phóng
        import gc
        gc.collect()
    
    # Kết hợp các lô đã xử lý
    logging.info(f"Kết hợp {len(result_dfs)} lô đã xử lý")
    result_df = pd.concat(result_dfs, ignore_index=True)
    
    return result_df

def simplify_html_report(report_data: Dict[str, Any], output_path: str) -> bool:
    """
    Tạo báo cáo HTML đơn giản dựa trên dữ liệu từ quá trình làm sạch
    
    Parameters:
    -----------
    report_data : Dict[str, Any]
        Dữ liệu báo cáo
    output_path : str
        Đường dẫn lưu file báo cáo HTML
        
    Returns:
    --------
    bool
        True nếu tạo báo cáo thành công, False nếu thất bại
    """
    try:
        # Tạo nội dung HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Báo cáo làm sạch dữ liệu</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1, h2, h3 {{ color: #444; }}
                .container {{ max-width: 1200px; margin: 0 auto; }}
                table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 6px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .summary {{ background-color: #f9f9f9; padding: 15px; border-radius: 4px; margin: 15px 0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Báo cáo làm sạch dữ liệu CSV</h1>
                <p><strong>Thời gian:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                
                <div class="summary">
                    <h2>Tóm tắt</h2>
                    <p><strong>File đầu vào:</strong> {report_data.get('input_file', 'N/A')}</p>
                    <p><strong>File đầu ra:</strong> {report_data.get('output_file', 'N/A')}</p>
                    <p><strong>File log:</strong> {report_data.get('log_file', 'N/A')}</p>
                </div>
                
                <h2>Thống kê dữ liệu</h2>
                <table>
                    <tr>
                        <th>Chỉ số</th>
                        <th>Trước khi làm sạch</th>
                        <th>Sau khi làm sạch</th>
                        <th>Thay đổi</th>
                    </tr>
        """
        
        # Thêm thống kê dữ liệu
        if 'stats' in report_data and 'before' in report_data['stats'] and 'after' in report_data['stats']:
            before = report_data['stats']['before']
            after = report_data['stats']['after']
            
            # Số dòng
            before_rows = before.get('row_count', 0)
            after_rows = after.get('row_count', 0)
            row_diff = before_rows - after_rows
            row_percent = f"{(abs(row_diff)/before_rows*100):.1f}%" if before_rows > 0 else "0.0%"
            html_content += f"""
                <tr>
                    <td>Số dòng</td>
                    <td>{before_rows:,}</td>
                    <td>{after_rows:,}</td>
                    <td>{'-' if row_diff >= 0 else '+'}{abs(row_diff):,}</td>
                </tr>
            """
            
            # Số cột
            before_cols = before.get('column_count', 0)
            after_cols = after.get('column_count', 0)
            col_diff = before_cols - after_cols
            html_content += f"""
                <tr>
                    <td>Số cột</td>
                    <td>{before_cols}</td>
                    <td>{after_cols}</td>
                    <td>{'-' if col_diff >= 0 else '+'}{abs(col_diff)}</td>
                </tr>
            """
            
            # Giá trị thiếu
            before_missing = before.get('missing_values', 0)
            after_missing = after.get('missing_values', 0)
            missing_diff = before_missing - after_missing
            html_content += f"""
                <tr>
                    <td>Giá trị thiếu</td>
                    <td>{before_missing:,}</td>
                    <td>{after_missing:,}</td>
                    <td>{'-' if missing_diff >= 0 else '+'}{abs(missing_diff):,}</td>
                </tr>
            """
            
            # Dung lượng bộ nhớ
            before_memory = before.get('memory_usage', 0)
            after_memory = after.get('memory_usage', 0)
            memory_diff = before_memory - after_memory
            memory_percent = f"{(abs(memory_diff)/before_memory*100):.1f}%" if before_memory > 0 else "0.0%"
            html_content += f"""
                <tr>
                    <td>Dung lượng bộ nhớ</td>
                    <td>{before_memory:.2f} MB</td>
                    <td>{after_memory:.2f} MB</td>
                    <td>{'-' if memory_diff >= 0 else '+'}{abs(memory_diff):.2f} MB ({memory_percent})</td>
                </tr>
            """
        
        html_content += """
                </table>
                
                <h2>Các bước xử lý</h2>
                <table>
                    <tr>
                        <th>Bước</th>
                        <th>Phương pháp</th>
                        <th>Chi tiết</th>
                    </tr>
        """
        
        # Thêm các bước xử lý
        if 'processing_steps' in report_data:
            for idx, step in enumerate(report_data['processing_steps'], 1):
                method = step.get('method', 'N/A')
                details = step.get('details', 'N/A')
                html_content += f"""
                    <tr>
                        <td>{idx}. {step.get('step', 'Bước ' + str(idx))}</td>
                        <td>{method}</td>
                        <td>{details}</td>
                    </tr>
                """
        
        html_content += """
                </table>
            </div>
        </body>
        </html>
        """
        
        # Lưu nội dung vào file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logging.info(f"Đã tạo báo cáo HTML tại: {output_path}")
        return True
        
    except Exception as e:
        logging.error(f"Lỗi khi tạo báo cáo HTML: {str(e)}")
        return False

def export_json_report(report_data: Dict[str, Any], output_path: str) -> bool:
    """
    Xuất báo cáo dạng JSON.
    
    Parameters:
    -----------
    report_data : Dict[str, Any]
        Dữ liệu báo cáo
    output_path : str
        Đường dẫn đến file JSON đầu ra
        
    Returns:
    --------
    bool
        True nếu xuất báo cáo thành công, False nếu có lỗi
    """
    try:
        # Đảm bảo thư mục đầu ra tồn tại
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # Chuyển đổi các kiểu dữ liệu không thể serializable sang chuỗi
        def convert_to_serializable(obj):
            if isinstance(obj, (np.int64, np.int32, np.float64, np.float32)):
                return int(obj) if isinstance(obj, (np.int64, np.int32)) else float(obj)
            elif isinstance(obj, (datetime, pd.Timestamp)):
                return obj.isoformat()
            elif isinstance(obj, dict):
                return {k: convert_to_serializable(v) for k, v in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [convert_to_serializable(i) for i in obj]
            elif isinstance(obj, pd.Series):
                return convert_to_serializable(obj.to_dict())
            elif isinstance(obj, pd.DataFrame):
                return convert_to_serializable(obj.to_dict())
            elif pd.isna(obj):
                return None
            else:
                return str(obj)
                
        # Chuyển đổi dữ liệu báo cáo để có thể serialize
        serializable_data = convert_to_serializable(report_data)
        
        # Ghi dữ liệu JSON ra file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(serializable_data, f, ensure_ascii=False, indent=2)
        
        print(f"Đã xuất báo cáo JSON tại: {output_path}")
        logging.info(f"Đã xuất báo cáo JSON tại: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo JSON: {str(e)}")
        logging.error(f"Lỗi khi xuất báo cáo JSON: {str(e)}")
        return False

def export_to_excel(
    df: pd.DataFrame, 
    output_path: str,
    sheet_name: str = 'Sheet1',
    index: bool = False,
    include_stats: bool = True,
    auto_adjust_columns: bool = True,
    create_summary_sheet: bool = False,
    apply_formatting: bool = True
) -> bool:
    """
    Xuất DataFrame ra file Excel với các tùy chọn nâng cao.
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame cần xuất
    output_path : str
        Đường dẫn xuất file Excel (.xlsx)
    sheet_name : str, default='Sheet1'
        Tên sheet chứa dữ liệu
    index : bool, default=False
        Có xuất chỉ mục (index) hay không
    include_stats : bool, default=True
        Có thêm thống kê cơ bản vào phần đầu của sheet không
    auto_adjust_columns : bool, default=True
        Tự động điều chỉnh độ rộng cột
    create_summary_sheet : bool, default=False
        Tạo sheet tóm tắt thống kê nếu True
    apply_formatting : bool, default=True
        Áp dụng định dạng tự động cho các cột dựa vào kiểu dữ liệu
        
    Returns:
    --------
    bool
        True nếu xuất thành công, False nếu có lỗi
    """
    try:
        # Kiểm tra DataFrame
        if df is None or df.empty:
            logging.error("DataFrame trống hoặc None, không thể xuất ra Excel")
            return False
        
        # Đảm bảo đường dẫn có phần mở rộng .xlsx
        if not output_path.lower().endswith('.xlsx'):
            output_path += '.xlsx'
            
        # Tạo thư mục đầu ra nếu chưa tồn tại
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        
        # Kiểm tra xem có engine Excel phù hợp không
        try:
            import importlib
            if importlib.util.find_spec("openpyxl") is None:
                logging.warning("Thư viện openpyxl chưa được cài đặt. Đang cài đặt...")
                print("Thư viện openpyxl chưa được cài đặt. Đang cài đặt...")
                import sys
                import subprocess
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                logging.info("Đã cài đặt openpyxl thành công")
        except Exception as e:
            logging.warning(f"Không thể cài đặt openpyxl: {str(e)}")
        
        # Tạo ExcelWriter với openpyxl
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Xuất dữ liệu chính
            if include_stats and not df.empty:
                # Tạo DataFrame thống kê
                stats_df = pd.DataFrame([
                    ['Thời gian xuất', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    ['Số dòng', df.shape[0]],
                    ['Số cột', df.shape[1]],
                    ['Dung lượng', f"{df.memory_usage(deep=True).sum() / (1024 * 1024):.2f} MB"],
                    ['Giá trị thiếu', f"{df.isna().sum().sum()} ({df.isna().sum().sum() / (df.shape[0] * df.shape[1]) * 100:.2f}%)"],
                    ['Kiểu dữ liệu', ', '.join([f"{k}: {v}" for k, v in df.dtypes.value_counts().items()])]
                ], columns=['Thông tin', 'Giá trị'])
                
                # Ghi thống kê vào đầu sheet, sau đó là dữ liệu chính
                stats_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                df.to_excel(writer, sheet_name=sheet_name, index=index, startrow=len(stats_df) + 2, startcol=0)
                
                # Tính row offset do có thêm thống kê
                data_row_offset = len(stats_df) + 3  # +3 vì có header và khoảng trống
            else:
                # Chỉ xuất dữ liệu, không kèm thống kê
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                data_row_offset = 1  # Chỉ có header
            
            # Tạo sheet tóm tắt nếu được yêu cầu
            if create_summary_sheet and not df.empty:
                # Tạo summary sheet với các thống kê chi tiết
                # Thông tin chung
                summary_data = []
                summary_data.append(['THÔNG TIN CHUNG', ''])
                summary_data.append(['Thời gian xuất', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                summary_data.append(['Số dòng', df.shape[0]])
                summary_data.append(['Số cột', df.shape[1]])
                summary_data.append(['Dung lượng', f"{df.memory_usage(deep=True).sum() / (1024 * 1024):.2f} MB"])
                summary_data.append(['', ''])
                
                # Thống kê giá trị thiếu
                summary_data.append(['THỐNG KÊ GIÁ TRỊ THIẾU', ''])
                summary_data.append(['Tổng số giá trị thiếu', df.isna().sum().sum()])
                summary_data.append(['Tỷ lệ giá trị thiếu', f"{df.isna().sum().sum() / (df.shape[0] * df.shape[1]) * 100:.2f}%"])
                
                # Thêm thông tin về giá trị thiếu theo cột
                missing_by_col = df.isna().sum().sort_values(ascending=False)
                missing_by_col = missing_by_col[missing_by_col > 0]
                if not missing_by_col.empty:
                    summary_data.append(['', ''])
                    summary_data.append(['GIÁ TRỊ THIẾU THEO CỘT', ''])
                    for col, missing in missing_by_col.items():
                        summary_data.append([col, f"{missing} ({missing / len(df) * 100:.2f}%)"])
                
                # Tạo DataFrame từ dữ liệu tóm tắt
                summary_df = pd.DataFrame(summary_data, columns=['Thông tin', 'Giá trị'])
                summary_df.to_excel(writer, sheet_name='Tóm tắt', index=False)
            
            # Áp dụng định dạng tự động dựa trên kiểu dữ liệu
            if apply_formatting and not df.empty:
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Định dạng header - chữ đậm, nền xám nhạt
                for col_idx, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=data_row_offset, column=col_idx)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # Kiểm tra xem DataFrame có metadata định dạng không
                format_metadata = df.attrs.get('format_metadata', {})
                
                # Định dạng các cột dựa vào kiểu dữ liệu và metadata
                for col_idx, column in enumerate(df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    # Ưu tiên sử dụng định dạng từ metadata nếu có
                    if column in format_metadata:
                        number_format = format_metadata[column]
                        for row in range(data_row_offset + 1, data_row_offset + len(df) + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.number_format = number_format
                    # Nếu không có trong metadata, xác định dựa trên kiểu dữ liệu
                    else:
                        # Xác định kiểu dữ liệu của cột
                        if pd.api.types.is_numeric_dtype(df[column]):
                            # Định dạng số - phân tách hàng nghìn, 2 số thập phân cho float
                            if pd.api.types.is_integer_dtype(df[column]):
                                # Số nguyên - định dạng không có thập phân
                                number_format = '#,##0'
                            else:
                                # Số thực - định dạng 2 số thập phân
                                number_format = '#,##0.00'
                                
                            # Áp dụng định dạng số cho cột
                            for row in range(data_row_offset + 1, data_row_offset + len(df) + 1):
                                cell = worksheet.cell(row=row, column=col_idx)
                                cell.number_format = number_format
                                
                        elif pd.api.types.is_datetime64_any_dtype(df[column]):
                            # Kiểm tra xem có thông tin giờ không
                            has_time = False
                            try:
                                # Kiểm tra mẫu đầu tiên không phải NA
                                sample = df[column].dropna().iloc[0] if not df[column].dropna().empty else None
                                if sample and (sample.hour != 0 or sample.minute != 0 or sample.second != 0):
                                    has_time = True
                            except:
                                pass
                                
                            # Định dạng ngày tháng
                            date_format = 'DD/MM/YYYY HH:MM:SS' if has_time else 'DD/MM/YYYY'
                            for row in range(data_row_offset + 1, data_row_offset + len(df) + 1):
                                cell = worksheet.cell(row=row, column=col_idx)
                                cell.number_format = date_format
                                
                        else:
                            # Định dạng văn bản
                            for row in range(data_row_offset + 1, data_row_offset + len(df) + 1):
                                cell = worksheet.cell(row=row, column=col_idx)
                                cell.number_format = '@'  # Định dạng cho văn bản
            
            # Tự động điều chỉnh độ rộng cột
            if auto_adjust_columns:
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for idx, col in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = worksheet.cell(row=1, column=idx).column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column].width = adjusted_width
        
        file_size = os.path.getsize(output_path)
        print(f"Đã xuất dữ liệu ra file Excel: {output_path} ({file_size / 1024:.1f} KB)")
        logging.info(f"Đã xuất dữ liệu ra file Excel: {output_path} ({file_size / 1024:.1f} KB)")
        return True
        
    except ImportError as e:
        logging.error(f"Lỗi thư viện khi xuất Excel: {str(e)}")
        print(f"Lỗi thư viện: {str(e)}. Hãy cài đặt thư viện openpyxl: pip install openpyxl")
        return False
    except PermissionError:
        logging.error(f"Lỗi quyền truy cập khi xuất Excel: {output_path}")
        print(f"Lỗi quyền truy cập! File {output_path} có thể đang được mở trong ứng dụng khác.")
        return False
    except Exception as e:
        logging.error(f"Lỗi khi xuất Excel: {str(e)}")
        print(f"Lỗi khi xuất Excel: {str(e)}")
        return False
